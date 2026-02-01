import logging
from openpyxl import load_workbook
from app.models import TestCase, TestStep

logger = logging.getLogger(__name__)


class XlsxReader:
    def __init__(self, filepath, column_mapping):
        """
        Args:
            filepath: Path to the xlsx file
            column_mapping: dict like {"id": 0, "title": 1, "step_no": 5, "operation": 6}
                           Values are column indices (0-based).
        """
        self.filepath = filepath
        self.mapping = column_mapping

    def read_all(self):
        """Parse the entire xlsx into (TestCase, [TestStep]) groups.

        Returns:
            list of (TestCase, list[TestStep]) tuples
        """
        wb = load_workbook(self.filepath, data_only=True)
        ws = wb.active

        # Handle merged cells: fill values into all cells of merged ranges
        self._unmerge_and_fill(ws)

        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).strip() if cell.value is not None else "")

        # Determine extra column indices
        core_indices = set(self.mapping.values())
        extra_indices = {}
        for i, header in enumerate(headers):
            if i not in core_indices and header:
                extra_indices[header] = i

        id_col = self.mapping.get("id")
        title_col = self.mapping.get("title")
        step_no_col = self.mapping.get("step_no")
        operation_col = self.mapping.get("operation")

        cases = {}  # case_id -> (TestCase, [TestStep])
        current_case_id = None
        current_title = None
        current_case_extra = {}

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            values = [cell.value for cell in row]

            # Get cell values for core fields
            raw_id = self._get_cell(values, id_col)
            raw_title = self._get_cell(values, title_col)
            raw_step_no = self._get_cell(values, step_no_col)
            raw_operation = self._get_cell(values, operation_col)

            # Update current case if a new ID appears
            if raw_id is not None and str(raw_id).strip():
                current_case_id = str(raw_id).strip()
                current_title = str(raw_title).strip() if raw_title else ""
                # Collect case-level extra fields
                current_case_extra = {}
                for col_name, col_idx in extra_indices.items():
                    val = self._get_cell(values, col_idx)
                    if val is not None and str(val).strip():
                        current_case_extra[col_name] = str(val).strip()
            elif raw_title is not None and str(raw_title).strip():
                # Title without ID - update title if we have a current case
                if current_case_id:
                    current_title = str(raw_title).strip()

            # Skip rows without a current case or without operation
            if not current_case_id:
                logger.debug("Skipping row %d: no case ID yet", row_idx)
                continue

            if raw_operation is None or not str(raw_operation).strip():
                logger.debug("Skipping row %d: empty operation", row_idx)
                continue

            # Parse step number
            step_no = 0
            if raw_step_no is not None:
                try:
                    step_no = int(float(str(raw_step_no).strip()))
                except (ValueError, TypeError):
                    step_no = 0

            operation = str(raw_operation).strip()

            # Collect step-level extra fields
            step_extra = {}
            for col_name, col_idx in extra_indices.items():
                val = self._get_cell(values, col_idx)
                if val is not None and str(val).strip():
                    step_extra[col_name] = str(val).strip()

            # Create or update case
            if current_case_id not in cases:
                case = TestCase(
                    id=current_case_id,
                    title=current_title or "",
                    extra_fields=current_case_extra,
                )
                cases[current_case_id] = (case, [])
            else:
                # Update title if it was empty before
                existing_case = cases[current_case_id][0]
                if not existing_case.title and current_title:
                    existing_case.title = current_title

            step = TestStep(
                case_id=current_case_id,
                step_no=step_no,
                operation=operation,
                extra_fields=step_extra,
            )
            cases[current_case_id][1].append(step)

        wb.close()

        result = list(cases.values())
        # Auto-number steps if step_no is 0
        for case, steps in result:
            all_zero = all(s.step_no == 0 for s in steps)
            if all_zero:
                for i, s in enumerate(steps, 1):
                    s.step_no = i

        total_cases = len(result)
        total_steps = sum(len(steps) for _, steps in result)
        logger.info("Parsed %d cases with %d steps from %s", total_cases, total_steps, self.filepath)

        return result

    def _unmerge_and_fill(self, ws):
        """Handle merged cells by filling the merged value to all cells in the range."""
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(str(merged_range))
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col, value=top_left_value)

    @staticmethod
    def _get_cell(values, col_idx):
        """Safely get a cell value by index."""
        if col_idx is None or col_idx >= len(values):
            return None
        val = values[col_idx]
        if hasattr(val, 'value'):
            return val.value
        return val
