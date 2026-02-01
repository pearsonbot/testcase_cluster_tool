import logging
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Each required field has a list of known aliases (case-insensitive matching)
FIELD_ALIASES = {
    "id": ["标识", "用例标识", "编号", "用例编号", "ID", "CaseID", "case_id", "Case ID"],
    "title": ["标题", "用例标题", "名称", "用例名称", "Title", "Case Title"],
    "step_no": ["TC步骤", "步骤", "步骤号", "步骤序号", "Step", "StepNo", "Step No"],
    "operation": ["TC操作", "操作", "步骤操作", "操作描述", "Operation", "Action", "步骤描述"],
}


class ColumnMapper:
    def __init__(self, filepath):
        self.filepath = filepath
        self.headers = []
        self.mapping = {}
        self.extra_columns = {}
        self._read_headers()

    def _read_headers(self):
        """Read the first row of the xlsx as headers."""
        wb = load_workbook(self.filepath, read_only=True, data_only=True)
        ws = wb.active
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        wb.close()

        if first_row is None:
            raise ValueError("The xlsx file is empty or has no header row")

        self.headers = [str(cell).strip() if cell is not None else "" for cell in first_row]
        logger.debug("Headers found: %s", self.headers)

    def auto_detect(self):
        """Auto-detect column mapping.

        Returns:
            tuple: (mapping_dict, list_of_unmatched_required_fields)
                mapping_dict: {field_name: column_index}
                unmatched: list of field names that could not be matched
        """
        self.mapping = {}
        used_indices = set()

        # Pass 1: Exact match (case-insensitive)
        for field_name, aliases in FIELD_ALIASES.items():
            if field_name in self.mapping:
                continue
            for i, header in enumerate(self.headers):
                if i in used_indices:
                    continue
                header_lower = header.lower().strip()
                for alias in aliases:
                    if header_lower == alias.lower():
                        self.mapping[field_name] = i
                        used_indices.add(i)
                        logger.info("Column '%s' (index %d) matched to field '%s' (exact)", header, i, field_name)
                        break
                if field_name in self.mapping:
                    break

        # Pass 2: Substring/contains match
        for field_name, aliases in FIELD_ALIASES.items():
            if field_name in self.mapping:
                continue
            for i, header in enumerate(self.headers):
                if i in used_indices:
                    continue
                header_lower = header.lower().strip()
                for alias in aliases:
                    if alias.lower() in header_lower or header_lower in alias.lower():
                        self.mapping[field_name] = i
                        used_indices.add(i)
                        logger.info("Column '%s' (index %d) matched to field '%s' (substring)", header, i, field_name)
                        break
                if field_name in self.mapping:
                    break

        # Collect extra columns
        self.extra_columns = {}
        for i, header in enumerate(self.headers):
            if i not in used_indices and header:
                self.extra_columns[header] = i

        unmatched = [f for f in FIELD_ALIASES if f not in self.mapping]
        if unmatched:
            logger.warning("Unmatched required fields: %s", unmatched)

        return self.mapping, unmatched
