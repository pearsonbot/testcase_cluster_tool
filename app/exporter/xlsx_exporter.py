import io
import re
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


def _sanitize_sheet_name(name, max_len=31):
    """Sanitize a string for use as an xlsx sheet name."""
    # Remove invalid characters
    name = re.sub(r'[\\/*?\[\]:]', '', name)
    if len(name) > max_len:
        name = name[:max_len - 3] + "..."
    return name or "Sheet"


def _style_header_row(ws, col_count):
    """Apply header styling to the first row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def _auto_width(ws):
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                # Estimate width for Chinese characters (2x width)
                val = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_length = max(max_length, length)
        adjusted = min(max_length + 4, 60)
        ws.column_dimensions[col_letter].width = adjusted


class XlsxExporter:
    def __init__(self, db):
        self.db = db

    def export_overview(self):
        """Generate cluster overview xlsx. Returns bytes."""
        wb = Workbook()
        ws = wb.active
        ws.title = "聚类总览"

        headers = ["簇编号", "簇标签", "步骤数量", "涉及用例数"]
        ws.append(headers)
        _style_header_row(ws, len(headers))

        rows = self.db.execute(
            "SELECT cluster_id, label, step_count, case_count "
            "FROM cluster_info ORDER BY cluster_id"
        ).fetchall()

        for r in rows:
            ws.append([r['cluster_id'], r['label'], r['step_count'], r['case_count']])

        _auto_width(ws)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def export_cluster_details(self):
        """Generate cluster details xlsx with one sheet per cluster. Returns bytes."""
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        clusters = self.db.execute(
            "SELECT cluster_id, label FROM cluster_info ORDER BY cluster_id"
        ).fetchall()

        if not clusters:
            ws = wb.create_sheet("无数据")
            ws.append(["No clustering data available"])
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        for cluster in clusters:
            cid = cluster['cluster_id']
            label = cluster['label'] or f"Cluster {cid}"
            sheet_name = _sanitize_sheet_name(f"簇{cid}_{label}")

            ws = wb.create_sheet(title=sheet_name)
            headers = ["步骤操作", "所属用例标识", "所属用例标题", "步骤号"]
            ws.append(headers)
            _style_header_row(ws, len(headers))

            steps = self.db.execute(
                "SELECT ts.operation, ts.case_id, tc.title, ts.step_no "
                "FROM cluster_results cr "
                "JOIN test_steps ts ON cr.step_id = ts.id "
                "JOIN test_cases tc ON ts.case_id = tc.id "
                "WHERE cr.cluster_id = ? "
                "ORDER BY ts.case_id, ts.step_no",
                (cid,)
            ).fetchall()

            for s in steps:
                ws.append([s['operation'], s['case_id'], s['title'], s['step_no']])

            _auto_width(ws)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def export_case_cluster_view(self):
        """Generate case-level cluster view xlsx. Returns bytes."""
        wb = Workbook()
        ws = wb.active
        ws.title = "用例聚类视图"

        headers = ["用例标识", "用例标题", "步骤号", "步骤操作", "簇编号", "簇标签"]
        ws.append(headers)
        _style_header_row(ws, len(headers))

        rows = self.db.execute(
            "SELECT tc.id as case_id, tc.title, ts.step_no, ts.operation, "
            "cr.cluster_id, cr.cluster_label "
            "FROM test_steps ts "
            "JOIN test_cases tc ON ts.case_id = tc.id "
            "LEFT JOIN cluster_results cr ON ts.id = cr.step_id "
            "ORDER BY tc.id, ts.step_no"
        ).fetchall()

        for r in rows:
            cluster_id = r['cluster_id'] if r['cluster_id'] is not None and r['cluster_id'] >= 0 else ""
            cluster_label = r['cluster_label'] or ""
            if r['cluster_id'] is not None and r['cluster_id'] < 0:
                cluster_label = "(independent)"

            ws.append([
                r['case_id'], r['title'], r['step_no'],
                r['operation'], cluster_id, cluster_label
            ])

        _auto_width(ws)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
